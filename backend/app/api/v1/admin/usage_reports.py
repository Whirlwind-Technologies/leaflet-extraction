"""
Admin API endpoints for Organization VLM Usage Reporting.

These endpoints provide comprehensive usage reports for organizations,
allowing super admins to monitor VLM consumption, costs, and trends.

Features:
- Organization-wise usage reports
- Time-based filtering and aggregation
- Cost analysis and trends
- Provider-specific usage breakdowns
- Export functionality (CSV, Excel, JSON)
"""

from typing import List, Optional, Dict, Any
from datetime import datetime, timedelta
from uuid import UUID
from decimal import Decimal
import csv
import io
from tempfile import NamedTemporaryFile

from fastapi import APIRouter, Depends, HTTPException, status, Query, Response, Body
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_, desc

from app.api.deps import get_db, get_current_superuser
from app.models.user import User
from app.models.organization import Organization
from app.models.organization_user import OrganizationUser
from app.models.organization_usage import OrganizationVLMUsage, OrganizationUsageSummary
from app.models.platform_vlm_provider import PlatformVLMProvider
from app.services.budget_monitoring_service import BudgetMonitoringService
from app.services.vlm_audit_service import VLMAuditService
from app.schemas.platform_vlm import (
    OrganizationUsageFilter,
    OrganizationUsageReport,
    UsageReportResponse
)

router = APIRouter()


class ExportRequest(BaseModel):
    """Request schema for exporting usage reports."""
    format: str = "csv"
    organization_id: Optional[str] = None
    provider_id: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    period_type: Optional[str] = None


@router.get("/filter-options")
async def get_filter_options(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_superuser),
) -> Dict[str, Any]:
    """
    Get filter options for usage reports - only returns organizations and providers
    that have usage data.
    """
    # Get organizations with usage data
    org_query = (
        select(Organization.id, Organization.name)
        .join(OrganizationVLMUsage, Organization.id == OrganizationVLMUsage.organization_id)
        .distinct()
        .order_by(Organization.name)
    )
    org_result = await db.execute(org_query)
    organizations = [
        {"id": str(row.id), "name": row.name}
        for row in org_result.all()
    ]

    # Get providers with usage data
    provider_query = (
        select(PlatformVLMProvider.id, PlatformVLMProvider.name, PlatformVLMProvider.provider_type)
        .join(OrganizationVLMUsage, PlatformVLMProvider.id == OrganizationVLMUsage.platform_provider_id)
        .distinct()
        .order_by(PlatformVLMProvider.name)
    )
    provider_result = await db.execute(provider_query)
    providers = [
        {
            "id": str(row.id),
            "name": row.name,
            "provider_type": row.provider_type.value if row.provider_type else None
        }
        for row in provider_result.all()
    ]

    return {
        "organizations": organizations,
        "providers": providers,
    }


@router.get("")
async def get_usage_reports(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_superuser),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    organization_id: Optional[UUID] = Query(None),
    provider_id: Optional[UUID] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
) -> Dict[str, Any]:
    """
    Get paginated usage reports.

    This is the main endpoint the frontend uses.
    """
    audit_service = VLMAuditService(db)

    # Parse dates - handle both date strings (YYYY-MM-DD) and datetime strings
    date_to = None
    date_from = None

    if end_date:
        try:
            # Check if it's a date-only string (YYYY-MM-DD) vs full datetime
            if len(end_date) == 10 and end_date[4] == '-' and end_date[7] == '-':
                # Date-only: set to end of day
                date_to = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            else:
                # Full datetime string
                date_to = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
        except ValueError:
            date_to = None

    if start_date:
        try:
            # Check if it's a date-only string (YYYY-MM-DD) vs full datetime
            if len(start_date) == 10 and start_date[4] == '-' and start_date[7] == '-':
                # Date-only: set to start of day
                date_from = datetime.strptime(start_date, "%Y-%m-%d")
            else:
                # Full datetime string
                date_from = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        except ValueError:
            date_from = None

    # Default date range: last 30 days
    if date_to is None:
        date_to = datetime.utcnow()
    if date_from is None:
        date_from = date_to - timedelta(days=30)

    try:
        # Build query
        query = select(OrganizationVLMUsage).where(
            and_(
                OrganizationVLMUsage.usage_date >= date_from,
                OrganizationVLMUsage.usage_date <= date_to
            )
        )

        if organization_id:
            query = query.where(OrganizationVLMUsage.organization_id == organization_id)
        if provider_id:
            query = query.where(OrganizationVLMUsage.platform_provider_id == provider_id)

        # Get total count
        count_query = select(func.count()).select_from(query.subquery())
        total_result = await db.execute(count_query)
        total = total_result.scalar() or 0

        # Apply pagination and ordering
        query = query.order_by(desc(OrganizationVLMUsage.usage_date))
        query = query.offset(skip).limit(limit)

        result = await db.execute(query)
        usage_records = result.scalars().all()

        # Get organization and provider names
        reports = []
        org_cache = {}
        provider_cache = {}

        for record in usage_records:
            # Cache organization lookup
            if record.organization_id not in org_cache:
                org_result = await db.execute(
                    select(Organization).where(Organization.id == record.organization_id)
                )
                org = org_result.scalar_one_or_none()
                org_cache[record.organization_id] = org.name if org else "Unknown"

            # Cache provider lookup
            if record.platform_provider_id and record.platform_provider_id not in provider_cache:
                provider_result = await db.execute(
                    select(PlatformVLMProvider).where(PlatformVLMProvider.id == record.platform_provider_id)
                )
                provider = provider_result.scalar_one_or_none()
                provider_cache[record.platform_provider_id] = {
                    "name": provider.name if provider else "Unknown",
                    "type": provider.provider_type.value if provider and provider.provider_type else None
                }

            provider_info = provider_cache.get(record.platform_provider_id, {"name": "Unknown", "type": None})

            reports.append({
                "id": str(record.id),
                "organization_id": str(record.organization_id),
                "organization_name": org_cache.get(record.organization_id, "Unknown"),
                "platform_provider_id": str(record.platform_provider_id) if record.platform_provider_id else None,
                "provider_name": provider_info["name"] if record.platform_provider_id else None,
                "provider_type": provider_info["type"] if record.platform_provider_id else None,
                "period_start": record.usage_date.isoformat(),
                "period_end": record.usage_date.isoformat(),
                "total_requests": record.request_count,
                "total_cost": float(record.total_cost) if record.total_cost else 0,
                "total_input_tokens": record.input_tokens,
                "total_output_tokens": record.output_tokens,
                # OrganizationVLMUsage does not yet track response times,
                # success/failure counts, or error counts. Returning null
                # (instead of a misleading 100.0 / 0) lets the dashboard
                # display "N/A" until those metrics are wired up.
                "average_response_time": None,
                "success_rate": None,
                "avg_cost_per_request": (float(record.total_cost) / record.request_count) if record.request_count > 0 else 0,
                "avg_tokens_per_request": ((record.input_tokens + record.output_tokens) / record.request_count) if record.request_count > 0 else 0,
                "error_count": None,
                "created_at": record.created_at.isoformat() if record.created_at else None,
            })

        audit_service.log_admin_access(
            admin_user_id=current_user.id,
            action="get_usage_reports",
            resource_type="usage_reports",
            filters={
                "organization_id": str(organization_id) if organization_id else None,
                "provider_id": str(provider_id) if provider_id else None,
                "start_date": date_from.isoformat(),
                "end_date": date_to.isoformat(),
            }
        )

        return {
            "items": reports,
            "total": total,
            "skip": skip,
            "limit": limit,
        }

    except Exception as e:
        audit_service.log_admin_error(
            admin_user_id=current_user.id,
            action="get_usage_reports",
            error_message=str(e)
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to get usage reports: {str(e)}"
        )


@router.get("/analytics")
async def get_usage_analytics(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_superuser),
    organization_id: Optional[UUID] = Query(None),
    provider_id: Optional[UUID] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
) -> Dict[str, Any]:
    """
    Get usage analytics summary for the admin dashboard.
    """
    audit_service = VLMAuditService(db)

    # Parse dates
    if end_date:
        try:
            date_to = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
        except:
            date_to = datetime.utcnow()
    else:
        date_to = datetime.utcnow()

    if start_date:
        try:
            date_from = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        except:
            date_from = date_to - timedelta(days=30)
    else:
        date_from = date_to - timedelta(days=30)

    try:
        # Build query
        query = select(OrganizationVLMUsage).where(
            and_(
                OrganizationVLMUsage.usage_date >= date_from,
                OrganizationVLMUsage.usage_date <= date_to
            )
        )

        if organization_id:
            query = query.where(OrganizationVLMUsage.organization_id == organization_id)
        if provider_id:
            query = query.where(OrganizationVLMUsage.platform_provider_id == provider_id)

        result = await db.execute(query)
        usage_records = result.scalars().all()

        # Calculate totals
        total_requests = sum(r.request_count for r in usage_records)
        total_cost = sum(float(r.total_cost) if r.total_cost else 0 for r in usage_records)
        total_input_tokens = sum(r.input_tokens for r in usage_records)
        total_output_tokens = sum(r.output_tokens for r in usage_records)

        # Get unique organizations
        org_ids = set(r.organization_id for r in usage_records)

        # Get actual unique users count from these organizations
        unique_users_count = 0
        if org_ids:
            users_query = select(func.count(func.distinct(OrganizationUser.user_id))).where(
                and_(
                    OrganizationUser.organization_id.in_(list(org_ids)),
                    OrganizationUser.is_active == True
                )
            )
            users_result = await db.execute(users_query)
            unique_users_count = users_result.scalar() or 0

        # Provider breakdown
        provider_costs = {}
        for r in usage_records:
            if r.platform_provider_id:
                pid = str(r.platform_provider_id)
                if pid not in provider_costs:
                    provider_costs[pid] = {"cost": 0, "requests": 0}
                provider_costs[pid]["cost"] += float(r.total_cost) if r.total_cost else 0
                provider_costs[pid]["requests"] += r.request_count

        # Get provider names
        cost_breakdown = []
        for pid, data in provider_costs.items():
            try:
                provider_result = await db.execute(
                    select(PlatformVLMProvider).where(PlatformVLMProvider.id == UUID(pid))
                )
                provider = provider_result.scalar_one_or_none()
                provider_name = provider.name if provider else f"Provider {pid[:8]}"
            except:
                provider_name = f"Provider {pid[:8]}"

            cost_breakdown.append({
                "provider": provider_name,
                "cost": data["cost"],
                "percentage": (data["cost"] / total_cost * 100) if total_cost > 0 else 0
            })

        # Trends data (by day)
        trends_by_date = {}
        for r in usage_records:
            date_key = r.usage_date.strftime("%Y-%m-%d") if hasattr(r.usage_date, 'strftime') else str(r.usage_date)[:10]
            if date_key not in trends_by_date:
                trends_by_date[date_key] = {"requests": 0, "cost": 0}
            trends_by_date[date_key]["requests"] += r.request_count
            trends_by_date[date_key]["cost"] += float(r.total_cost) if r.total_cost else 0

        trends_data = [
            {"date": date, "requests": data["requests"], "cost": data["cost"]}
            for date, data in sorted(trends_by_date.items())
        ]

        audit_service.log_admin_access(
            admin_user_id=current_user.id,
            action="get_usage_analytics",
            resource_type="usage_analytics",
        )

        # Find top provider (by request count, not cost)
        top_provider = {"name": "N/A", "requests": 0, "percentage": 0}
        if provider_costs:
            # Find provider with most requests
            top_pid = max(provider_costs.keys(), key=lambda pid: provider_costs[pid]["requests"])
            top_data = provider_costs[top_pid]

            # Get provider name
            try:
                provider_result = await db.execute(
                    select(PlatformVLMProvider).where(PlatformVLMProvider.id == UUID(top_pid))
                )
                provider = provider_result.scalar_one_or_none()
                provider_name = provider.name if provider else f"Provider {top_pid[:8]}"
            except:
                provider_name = f"Provider {top_pid[:8]}"

            request_percentage = (top_data["requests"] / total_requests * 100) if total_requests > 0 else 0
            top_provider = {
                "name": provider_name,
                "requests": top_data["requests"],
                "percentage": round(request_percentage, 1)
            }

        # Determine cost trend
        cost_trend = "stable"
        if len(trends_data) >= 2:
            recent_cost = sum(t["cost"] for t in trends_data[-7:]) if len(trends_data) >= 7 else sum(t["cost"] for t in trends_data[-len(trends_data)//2:])
            older_cost = sum(t["cost"] for t in trends_data[:7]) if len(trends_data) >= 7 else sum(t["cost"] for t in trends_data[:len(trends_data)//2])
            if older_cost > 0:
                change_pct = ((recent_cost - older_cost) / older_cost) * 100
                if change_pct > 10:
                    cost_trend = "increasing"
                elif change_pct < -10:
                    cost_trend = "decreasing"

        return {
            "total_requests": total_requests,
            "total_cost": total_cost,
            "total_input_tokens": total_input_tokens,
            "total_output_tokens": total_output_tokens,
            "organization_count": len(org_ids),
            "unique_organizations": len(org_ids),
            "unique_users": unique_users_count,
            "avg_cost_per_request": total_cost / total_requests if total_requests > 0 else 0,
            # Real success/failure is not tracked in OrganizationVLMUsage yet;
            # surface null so the dashboard can show "N/A" instead of a
            # misleading 100.0.
            "success_rate": None,
            "top_provider": top_provider,
            "cost_trend": cost_trend,
            "trends_data": trends_data,
            "cost_breakdown": cost_breakdown,
        }

    except Exception as e:
        audit_service.log_admin_error(
            admin_user_id=current_user.id,
            action="get_usage_analytics",
            error_message=str(e)
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to get usage analytics: {str(e)}"
        )


@router.get("/organizations", response_model=UsageReportResponse)
async def get_organization_usage_reports(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_superuser),
    organization_ids: Optional[List[UUID]] = Query(None, description="Filter by organization IDs"),
    provider_ids: Optional[List[UUID]] = Query(None, description="Filter by platform provider IDs"),
    days: int = Query(30, ge=1, le=365, description="Number of days for report"),
    date_from: Optional[datetime] = Query(None, description="Start date (overrides days)"),
    date_to: Optional[datetime] = Query(None, description="End date (overrides days)"),
    min_cost: Optional[Decimal] = Query(None, ge=0, description="Minimum cost threshold"),
    include_inactive: bool = Query(False, description="Include inactive organizations"),
) -> UsageReportResponse:
    """
    Generate comprehensive usage reports for organizations.

    Provides detailed VLM usage analytics including costs, token consumption,
    and provider breakdowns for specified organizations and time periods.
    """
    budget_service = BudgetMonitoringService(db)
    audit_service = VLMAuditService(db)

    # Determine date range
    if date_to is None:
        date_to = datetime.utcnow()
    if date_from is None:
        date_from = date_to - timedelta(days=days)

    # Log admin access
    audit_service.log_admin_access(
        admin_user_id=current_user.id,
        action="get_organization_usage_reports",
        resource_type="usage_reports",
        filters={
            "organization_ids": [str(id) for id in organization_ids] if organization_ids else None,
            "provider_ids": [str(id) for id in provider_ids] if provider_ids else None,
            "date_from": date_from.isoformat(),
            "date_to": date_to.isoformat(),
            "min_cost": float(min_cost) if min_cost else None,
        }
    )

    try:
        # Build base query for organizations
        org_query = select(Organization)
        if not include_inactive:
            org_query = org_query.where(Organization.is_active == True)
        if organization_ids:
            org_query = org_query.where(Organization.id.in_(organization_ids))

        org_result = await db.execute(org_query)
        organizations = org_result.scalars().all()

        reports = []
        total_cost = Decimal(0)
        total_requests = 0
        total_input_tokens = 0
        total_output_tokens = 0

        for org in organizations:
            # Get usage data for this organization
            usage_query = select(OrganizationVLMUsage).where(
                and_(
                    OrganizationVLMUsage.organization_id == org.id,
                    OrganizationVLMUsage.usage_date >= date_from,
                    OrganizationVLMUsage.usage_date <= date_to
                )
            )

            if provider_ids:
                usage_query = usage_query.where(
                    OrganizationVLMUsage.platform_provider_id.in_(provider_ids)
                )

            usage_result = await db.execute(usage_query)
            usage_records = usage_result.scalars().all()

            if not usage_records:
                continue

            # Aggregate organization totals
            org_requests = sum(r.request_count for r in usage_records)
            org_input_tokens = sum(r.input_tokens for r in usage_records)
            org_output_tokens = sum(r.output_tokens for r in usage_records)
            org_cost = sum(r.cost for r in usage_records)

            # Apply minimum cost filter
            if min_cost and org_cost < min_cost:
                continue

            # Get provider breakdown
            providers_used = {}
            for record in usage_records:
                provider_id = record.platform_provider_id
                if provider_id not in providers_used:
                    provider_query = select(PlatformVLMProvider).where(
                        PlatformVLMProvider.id == provider_id
                    )
                    provider_result = await db.execute(provider_query)
                    provider = provider_result.scalar_one_or_none()

                    providers_used[provider_id] = {
                        "provider_id": str(provider_id),
                        "provider_name": provider.name if provider else "Unknown",
                        "provider_type": provider.provider_type.value if provider else "unknown",
                        "model_name": provider.model_name if provider else "unknown",
                        "requests": 0,
                        "input_tokens": 0,
                        "output_tokens": 0,
                        "cost": Decimal(0)
                    }

                providers_used[provider_id]["requests"] += record.request_count
                providers_used[provider_id]["input_tokens"] += record.input_tokens
                providers_used[provider_id]["output_tokens"] += record.output_tokens
                providers_used[provider_id]["cost"] += record.cost

            # Get daily breakdown
            daily_breakdown = {}
            for record in usage_records:
                date_key = record.usage_date.strftime("%Y-%m-%d")
                if date_key not in daily_breakdown:
                    daily_breakdown[date_key] = {
                        "date": date_key,
                        "requests": 0,
                        "input_tokens": 0,
                        "output_tokens": 0,
                        "cost": Decimal(0),
                        "avg_cost_per_request": Decimal(0)
                    }

                daily_breakdown[date_key]["requests"] += record.request_count
                daily_breakdown[date_key]["input_tokens"] += record.input_tokens
                daily_breakdown[date_key]["output_tokens"] += record.output_tokens
                daily_breakdown[date_key]["cost"] += record.cost

            # Calculate averages for daily breakdown
            for day_data in daily_breakdown.values():
                if day_data["requests"] > 0:
                    day_data["avg_cost_per_request"] = day_data["cost"] / day_data["requests"]

            # Create organization report
            report = OrganizationUsageReport(
                organization_id=org.id,
                organization_name=org.name,
                period_start=date_from,
                period_end=date_to,
                total_requests=org_requests,
                total_input_tokens=org_input_tokens,
                total_output_tokens=org_output_tokens,
                total_cost=org_cost,
                average_cost_per_request=org_cost / org_requests if org_requests > 0 else Decimal(0),
                providers_used=list(providers_used.values()),
                daily_breakdown=list(daily_breakdown.values())
            )

            reports.append(report)

            # Add to totals
            total_cost += org_cost
            total_requests += org_requests
            total_input_tokens += org_input_tokens
            total_output_tokens += org_output_tokens

        # Create summary
        summary = {
            "total_organizations": len(reports),
            "total_requests": total_requests,
            "total_input_tokens": total_input_tokens,
            "total_output_tokens": total_output_tokens,
            "total_cost": float(total_cost),
            "average_cost_per_request": float(total_cost / total_requests) if total_requests > 0 else 0,
            "average_cost_per_organization": float(total_cost / len(reports)) if reports else 0,
            "period_days": (date_to - date_from).days,
            "date_range": {
                "from": date_from.isoformat(),
                "to": date_to.isoformat()
            }
        }

        # Sort reports by total cost (highest first)
        reports.sort(key=lambda r: r.total_cost, reverse=True)

        return UsageReportResponse(
            reports=reports,
            summary=summary,
            generated_at=datetime.utcnow(),
            period_days=(date_to - date_from).days,
            total_organizations=len(reports)
        )

    except Exception as e:
        audit_service.log_admin_error(
            admin_user_id=current_user.id,
            action="get_organization_usage_reports",
            error_message=str(e)
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate usage reports: {str(e)}"
        )


@router.get("/organizations/{organization_id}/detailed", response_model=OrganizationUsageReport)
async def get_detailed_organization_usage(
    organization_id: UUID,
    days: int = Query(30, ge=1, le=365, description="Number of days for report"),
    include_hourly: bool = Query(False, description="Include hourly breakdown"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_superuser),
) -> OrganizationUsageReport:
    """
    Get detailed usage report for a specific organization.

    Provides granular usage data including hourly breakdowns if requested.
    """
    audit_service = VLMAuditService(db)

    # Verify organization exists
    org_query = select(Organization).where(Organization.id == organization_id)
    org_result = await db.execute(org_query)
    org = org_result.scalar_one_or_none()

    if not org:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Organization not found"
        )

    date_to = datetime.utcnow()
    date_from = date_to - timedelta(days=days)

    # Log admin access
    audit_service.log_admin_access(
        admin_user_id=current_user.id,
        action="get_detailed_organization_usage",
        resource_type="organization_usage",
        resource_id=organization_id,
        filters={
            "days": days,
            "include_hourly": include_hourly,
        }
    )

    try:
        # Get usage data
        usage_query = select(OrganizationVLMUsage).where(
            and_(
                OrganizationVLMUsage.organization_id == organization_id,
                OrganizationVLMUsage.usage_date >= date_from,
                OrganizationVLMUsage.usage_date <= date_to
            )
        ).order_by(OrganizationVLMUsage.usage_date.desc())

        usage_result = await db.execute(usage_query)
        usage_records = usage_result.scalars().all()

        # Aggregate totals
        total_requests = sum(r.request_count for r in usage_records)
        total_input_tokens = sum(r.input_tokens for r in usage_records)
        total_output_tokens = sum(r.output_tokens for r in usage_records)
        total_cost = sum(r.cost for r in usage_records)

        # Provider breakdown
        providers_used = {}
        for record in usage_records:
            provider_id = record.platform_provider_id
            if provider_id not in providers_used:
                provider_query = select(PlatformVLMProvider).where(
                    PlatformVLMProvider.id == provider_id
                )
                provider_result = await db.execute(provider_query)
                provider = provider_result.scalar_one_or_none()

                providers_used[provider_id] = {
                    "provider_id": str(provider_id),
                    "provider_name": provider.name if provider else "Unknown",
                    "provider_type": provider.provider_type.value if provider else "unknown",
                    "model_name": provider.model_name if provider else "unknown",
                    "requests": 0,
                    "input_tokens": 0,
                    "output_tokens": 0,
                    "cost": Decimal(0),
                    "first_used": None,
                    "last_used": None
                }

            provider_data = providers_used[provider_id]
            provider_data["requests"] += record.request_count
            provider_data["input_tokens"] += record.input_tokens
            provider_data["output_tokens"] += record.output_tokens
            provider_data["cost"] += record.cost

            # Track usage dates
            usage_date = record.usage_date
            if not provider_data["first_used"] or usage_date < provider_data["first_used"]:
                provider_data["first_used"] = usage_date
            if not provider_data["last_used"] or usage_date > provider_data["last_used"]:
                provider_data["last_used"] = usage_date

        # Time breakdown (daily or hourly based on parameter)
        time_breakdown = {}
        if include_hourly:
            # Hourly breakdown for last 7 days only (to avoid huge response)
            recent_date_from = date_to - timedelta(days=7)
            for record in usage_records:
                if record.usage_date >= recent_date_from:
                    hour_key = record.usage_date.strftime("%Y-%m-%d %H:00")
                    if hour_key not in time_breakdown:
                        time_breakdown[hour_key] = {
                            "datetime": hour_key,
                            "requests": 0,
                            "input_tokens": 0,
                            "output_tokens": 0,
                            "cost": Decimal(0)
                        }

                    time_breakdown[hour_key]["requests"] += record.request_count
                    time_breakdown[hour_key]["input_tokens"] += record.input_tokens
                    time_breakdown[hour_key]["output_tokens"] += record.output_tokens
                    time_breakdown[hour_key]["cost"] += record.cost
        else:
            # Daily breakdown
            for record in usage_records:
                date_key = record.usage_date.strftime("%Y-%m-%d")
                if date_key not in time_breakdown:
                    time_breakdown[date_key] = {
                        "date": date_key,
                        "requests": 0,
                        "input_tokens": 0,
                        "output_tokens": 0,
                        "cost": Decimal(0)
                    }

                time_breakdown[date_key]["requests"] += record.request_count
                time_breakdown[date_key]["input_tokens"] += record.input_tokens
                time_breakdown[date_key]["output_tokens"] += record.output_tokens
                time_breakdown[date_key]["cost"] += record.cost

        return OrganizationUsageReport(
            organization_id=organization_id,
            organization_name=org.name,
            period_start=date_from,
            period_end=date_to,
            total_requests=total_requests,
            total_input_tokens=total_input_tokens,
            total_output_tokens=total_output_tokens,
            total_cost=total_cost,
            average_cost_per_request=total_cost / total_requests if total_requests > 0 else Decimal(0),
            providers_used=list(providers_used.values()),
            daily_breakdown=list(time_breakdown.values())
        )

    except Exception as e:
        audit_service.log_admin_error(
            admin_user_id=current_user.id,
            action="get_detailed_organization_usage",
            error_message=str(e),
            resource_id=organization_id
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate detailed usage report: {str(e)}"
        )


@router.post("/export")
async def export_usage_report(
    export_request: ExportRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_superuser),
) -> Dict[str, Any]:
    """
    Export usage reports in the specified format.

    Returns a download URL for the exported file.
    """
    import base64
    from datetime import timezone

    audit_service = VLMAuditService(db)

    format_type = export_request.format or "csv"
    organization_id = export_request.organization_id
    provider_id = export_request.provider_id
    start_date = export_request.start_date
    end_date = export_request.end_date

    # Parse dates
    if end_date:
        try:
            if len(end_date) == 10:
                date_to = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            else:
                date_to = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
        except ValueError:
            date_to = datetime.now(timezone.utc)
    else:
        date_to = datetime.now(timezone.utc)

    if start_date:
        try:
            if len(start_date) == 10:
                date_from = datetime.strptime(start_date, "%Y-%m-%d")
            else:
                date_from = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        except ValueError:
            date_from = date_to - timedelta(days=30)
    else:
        date_from = date_to - timedelta(days=30)

    try:
        # Build query
        query = select(OrganizationVLMUsage).where(
            and_(
                OrganizationVLMUsage.usage_date >= date_from,
                OrganizationVLMUsage.usage_date <= date_to
            )
        )

        if organization_id:
            query = query.where(OrganizationVLMUsage.organization_id == UUID(organization_id))
        if provider_id:
            query = query.where(OrganizationVLMUsage.platform_provider_id == UUID(provider_id))

        result = await db.execute(query.order_by(desc(OrganizationVLMUsage.usage_date)))
        usage_records = result.scalars().all()

        # Cache organization and provider names
        org_cache = {}
        provider_cache = {}

        for record in usage_records:
            if record.organization_id not in org_cache:
                org_result = await db.execute(
                    select(Organization).where(Organization.id == record.organization_id)
                )
                org = org_result.scalar_one_or_none()
                org_cache[record.organization_id] = org.name if org else "Unknown"

            if record.platform_provider_id and record.platform_provider_id not in provider_cache:
                provider_result = await db.execute(
                    select(PlatformVLMProvider).where(PlatformVLMProvider.id == record.platform_provider_id)
                )
                provider = provider_result.scalar_one_or_none()
                provider_cache[record.platform_provider_id] = {
                    "name": provider.name if provider else "Unknown",
                    "type": provider.provider_type.value if provider and provider.provider_type else None
                }

        # Build export data
        if format_type == "json":
            import json
            export_data = []
            for record in usage_records:
                provider_info = provider_cache.get(record.platform_provider_id, {"name": "Unknown", "type": None})
                export_data.append({
                    "organization_id": str(record.organization_id),
                    "organization_name": org_cache.get(record.organization_id, "Unknown"),
                    "provider_name": provider_info["name"],
                    "provider_type": provider_info["type"],
                    "date": record.usage_date.strftime("%Y-%m-%d"),
                    "requests": record.request_count,
                    "input_tokens": record.input_tokens,
                    "output_tokens": record.output_tokens,
                    "cost": float(record.total_cost) if record.total_cost else 0,
                })
            content = json.dumps(export_data, indent=2)
            content_type = "application/json"
            extension = "json"
            is_binary = False
        elif format_type == "excel":
            # Excel format using openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "Usage Report"

            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Write headers
            headers = ['Date', 'Organization', 'Provider', 'Provider Type',
                       'Requests', 'Input Tokens', 'Output Tokens', 'Cost ($)']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Write data rows
            for row_idx, record in enumerate(usage_records, 2):
                provider_info = provider_cache.get(record.platform_provider_id, {"name": "Unknown", "type": None})
                row_data = [
                    record.usage_date.strftime("%Y-%m-%d"),
                    org_cache.get(record.organization_id, "Unknown"),
                    provider_info["name"],
                    provider_info["type"] or "",
                    record.request_count,
                    record.input_tokens,
                    record.output_tokens,
                    float(record.total_cost) if record.total_cost else 0,
                ]
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.border = thin_border
                    if col == 8:  # Cost column
                        cell.number_format = '$#,##0.0000'

            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                max_length = len(str(headers[col - 1]))
                for row in range(2, len(usage_records) + 2):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 50)

            # Add summary sheet
            ws_summary = wb.create_sheet(title="Summary")
            summary_data = [
                ["Usage Report Summary", ""],
                ["", ""],
                ["Date Range", f"{date_from.strftime('%Y-%m-%d')} to {date_to.strftime('%Y-%m-%d')}"],
                ["Total Records", len(usage_records)],
                ["Total Requests", sum(r.request_count for r in usage_records)],
                ["Total Input Tokens", sum(r.input_tokens for r in usage_records)],
                ["Total Output Tokens", sum(r.output_tokens for r in usage_records)],
                ["Total Cost", sum(float(r.total_cost) if r.total_cost else 0 for r in usage_records)],
            ]
            for row_idx, (label, value) in enumerate(summary_data, 1):
                ws_summary.cell(row=row_idx, column=1, value=label).font = Font(bold=True) if row_idx == 1 else Font()
                ws_summary.cell(row=row_idx, column=2, value=value)
                if row_idx == 8:  # Total Cost row
                    ws_summary.cell(row=row_idx, column=2).number_format = '$#,##0.0000'

            ws_summary.column_dimensions['A'].width = 20
            ws_summary.column_dimensions['B'].width = 30

            # Save to bytes
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            content = excel_buffer.getvalue()
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            extension = "xlsx"
            is_binary = True
        else:
            # CSV format (default)
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow([
                'Date', 'Organization', 'Provider', 'Provider Type',
                'Requests', 'Input Tokens', 'Output Tokens', 'Cost'
            ])

            for record in usage_records:
                provider_info = provider_cache.get(record.platform_provider_id, {"name": "Unknown", "type": None})
                writer.writerow([
                    record.usage_date.strftime("%Y-%m-%d"),
                    org_cache.get(record.organization_id, "Unknown"),
                    provider_info["name"],
                    provider_info["type"] or "",
                    record.request_count,
                    record.input_tokens,
                    record.output_tokens,
                    float(record.total_cost) if record.total_cost else 0,
                ])

            content = output.getvalue()
            content_type = "text/csv"
            extension = "csv"
            is_binary = False

        # Create a data URL for download
        if is_binary:
            encoded_content = base64.b64encode(content).decode('utf-8')
        else:
            encoded_content = base64.b64encode(content.encode('utf-8')).decode('utf-8')
        download_url = f"data:{content_type};base64,{encoded_content}"

        # Log export
        audit_service.log_admin_action(
            admin_user_id=current_user.id,
            action="export_usage_report",
            resource_type="usage_reports",
            resource_data={
                "format": format_type,
                "record_count": len(usage_records),
                "date_range": f"{date_from.strftime('%Y-%m-%d')} to {date_to.strftime('%Y-%m-%d')}",
            }
        )

        return {
            "download_url": download_url,
            "expires_at": (datetime.now(timezone.utc) + timedelta(hours=1)).isoformat(),
            "filename": f"usage-report-{date_from.strftime('%Y%m%d')}-{date_to.strftime('%Y%m%d')}.{extension}",
            "record_count": len(usage_records),
        }

    except Exception as e:
        audit_service.log_admin_error(
            admin_user_id=current_user.id,
            action="export_usage_report",
            error_message=str(e)
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to export usage report: {str(e)}"
        )


@router.get("/export/csv")
async def export_usage_reports_csv(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_superuser),
    organization_ids: Optional[List[UUID]] = Query(None),
    days: int = Query(30, ge=1, le=365),
    include_daily_breakdown: bool = Query(False, description="Include daily breakdown in export"),
) -> StreamingResponse:
    """
    Export organization usage reports as CSV.

    Provides CSV download of usage data suitable for external analysis.
    """
    audit_service = VLMAuditService(db)

    try:
        # Get usage reports (reuse the main endpoint logic)
        date_to = datetime.utcnow()
        date_from = date_to - timedelta(days=days)

        # Build CSV content
        output = io.StringIO()
        writer = csv.writer(output)

        if include_daily_breakdown:
            # Detailed CSV with daily breakdown
            writer.writerow([
                'Organization ID', 'Organization Name', 'Date', 'Provider ID', 'Provider Name',
                'Provider Type', 'Model', 'Requests', 'Input Tokens', 'Output Tokens', 'Cost'
            ])

            # Get detailed data for CSV
            usage_query = select(OrganizationVLMUsage).where(
                and_(
                    OrganizationVLMUsage.usage_date >= date_from,
                    OrganizationVLMUsage.usage_date <= date_to
                )
            )
            if organization_ids:
                usage_query = usage_query.where(
                    OrganizationVLMUsage.organization_id.in_(organization_ids)
                )

            usage_result = await db.execute(usage_query)
            usage_records = usage_result.scalars().all()

            # Cache organizations and providers
            org_cache = {}
            provider_cache = {}

            for record in usage_records:
                # Get organization info
                if record.organization_id not in org_cache:
                    org_query = select(Organization).where(Organization.id == record.organization_id)
                    org_result = await db.execute(org_query)
                    org = org_result.scalar_one_or_none()
                    org_cache[record.organization_id] = org.name if org else "Unknown"

                # Get provider info
                if record.platform_provider_id not in provider_cache:
                    provider_query = select(PlatformVLMProvider).where(
                        PlatformVLMProvider.id == record.platform_provider_id
                    )
                    provider_result = await db.execute(provider_query)
                    provider = provider_result.scalar_one_or_none()
                    provider_cache[record.platform_provider_id] = {
                        "name": provider.name if provider else "Unknown",
                        "type": provider.provider_type.value if provider else "unknown",
                        "model": provider.model_name if provider else "unknown"
                    }

                provider_info = provider_cache[record.platform_provider_id]

                writer.writerow([
                    str(record.organization_id),
                    org_cache[record.organization_id],
                    record.usage_date.strftime("%Y-%m-%d"),
                    str(record.platform_provider_id),
                    provider_info["name"],
                    provider_info["type"],
                    provider_info["model"],
                    record.request_count,
                    record.input_tokens,
                    record.output_tokens,
                    float(record.cost)
                ])

        else:
            # Summary CSV
            writer.writerow([
                'Organization ID', 'Organization Name', 'Period Start', 'Period End',
                'Total Requests', 'Total Input Tokens', 'Total Output Tokens',
                'Total Cost', 'Avg Cost Per Request', 'Primary Provider'
            ])

            # Get summary data (reuse logic from main endpoint)
            # ... (implementation similar to get_organization_usage_reports but simplified for CSV)

        # Log export
        audit_service.log_admin_action(
            admin_user_id=current_user.id,
            action="export_usage_reports_csv",
            resource_type="usage_reports",
            resource_data={
                "organization_count": len(organization_ids) if organization_ids else "all",
                "days": days,
                "include_daily_breakdown": include_daily_breakdown,
            }
        )

        # Prepare streaming response
        output.seek(0)
        filename = f"vlm_usage_report_{date_from.strftime('%Y%m%d')}_{date_to.strftime('%Y%m%d')}.csv"

        return StreamingResponse(
            io.BytesIO(output.getvalue().encode('utf-8')),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        audit_service.log_admin_error(
            admin_user_id=current_user.id,
            action="export_usage_reports_csv",
            error_message=str(e)
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to export usage reports: {str(e)}"
        )


@router.get("/summary/trends")
async def get_usage_trends(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_superuser),
    days: int = Query(90, ge=7, le=365, description="Number of days for trend analysis"),
    group_by: str = Query("daily", description="Grouping: daily, weekly, monthly"),
) -> Dict[str, Any]:
    """
    Get usage trends and analytics across all organizations.

    Provides trend analysis for cost, usage patterns, and provider distribution.
    """
    audit_service = VLMAuditService(db)

    try:
        date_to = datetime.utcnow()
        date_from = date_to - timedelta(days=days)

        # Get all usage data for the period
        usage_query = select(OrganizationVLMUsage).where(
            and_(
                OrganizationVLMUsage.usage_date >= date_from,
                OrganizationVLMUsage.usage_date <= date_to
            )
        ).order_by(OrganizationVLMUsage.usage_date)

        usage_result = await db.execute(usage_query)
        usage_records = usage_result.scalars().all()

        # Group by time period
        trends = {}
        provider_trends = {}

        for record in usage_records:
            if group_by == "weekly":
                # Get Monday of the week
                days_since_monday = record.usage_date.weekday()
                period_start = record.usage_date - timedelta(days=days_since_monday)
                period_key = period_start.strftime("%Y-W%U")
            elif group_by == "monthly":
                period_key = record.usage_date.strftime("%Y-%m")
            else:  # daily
                period_key = record.usage_date.strftime("%Y-%m-%d")

            if period_key not in trends:
                trends[period_key] = {
                    "period": period_key,
                    "requests": 0,
                    "input_tokens": 0,
                    "output_tokens": 0,
                    "cost": 0,
                    "organizations": set(),
                    "providers": set()
                }

            trends[period_key]["requests"] += record.request_count
            trends[period_key]["input_tokens"] += record.input_tokens
            trends[period_key]["output_tokens"] += record.output_tokens
            trends[period_key]["cost"] += float(record.cost)
            trends[period_key]["organizations"].add(record.organization_id)
            trends[period_key]["providers"].add(record.platform_provider_id)

            # Provider-specific trends
            provider_id = str(record.platform_provider_id)
            if provider_id not in provider_trends:
                provider_trends[provider_id] = {}
            if period_key not in provider_trends[provider_id]:
                provider_trends[provider_id][period_key] = {
                    "requests": 0,
                    "cost": 0
                }
            provider_trends[provider_id][period_key]["requests"] += record.request_count
            provider_trends[provider_id][period_key]["cost"] += float(record.cost)

        # Convert sets to counts and sort by period
        trend_list = []
        for period_data in sorted(trends.values(), key=lambda x: x["period"]):
            period_data["organization_count"] = len(period_data["organizations"])
            period_data["provider_count"] = len(period_data["providers"])
            del period_data["organizations"]
            del period_data["providers"]
            trend_list.append(period_data)

        # Calculate growth rates
        for i in range(1, len(trend_list)):
            current = trend_list[i]
            previous = trend_list[i - 1]

            if previous["cost"] > 0:
                current["cost_growth_rate"] = ((current["cost"] - previous["cost"]) / previous["cost"]) * 100
            else:
                current["cost_growth_rate"] = 0 if current["cost"] == 0 else 100

            if previous["requests"] > 0:
                current["request_growth_rate"] = ((current["requests"] - previous["requests"]) / previous["requests"]) * 100
            else:
                current["request_growth_rate"] = 0 if current["requests"] == 0 else 100

        # Get provider names for trends
        provider_names = {}
        provider_query = select(PlatformVLMProvider)
        provider_result = await db.execute(provider_query)
        providers = provider_result.scalars().all()
        for provider in providers:
            provider_names[str(provider.id)] = provider.name

        # Log access
        audit_service.log_admin_access(
            admin_user_id=current_user.id,
            action="get_usage_trends",
            resource_type="usage_analytics",
            filters={
                "days": days,
                "group_by": group_by
            }
        )

        return {
            "period_type": group_by,
            "date_range": {
                "from": date_from.isoformat(),
                "to": date_to.isoformat()
            },
            "trends": trend_list,
            "provider_trends": {
                provider_names.get(pid, f"Provider {pid}"): periods
                for pid, periods in provider_trends.items()
            },
            "summary": {
                "total_periods": len(trend_list),
                "total_requests": sum(t["requests"] for t in trend_list),
                "total_cost": sum(t["cost"] for t in trend_list),
                "avg_requests_per_period": sum(t["requests"] for t in trend_list) / len(trend_list) if trend_list else 0,
                "avg_cost_per_period": sum(t["cost"] for t in trend_list) / len(trend_list) if trend_list else 0,
            },
            "generated_at": datetime.utcnow().isoformat()
        }

    except Exception as e:
        audit_service.log_admin_error(
            admin_user_id=current_user.id,
            action="get_usage_trends",
            error_message=str(e)
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate usage trends: {str(e)}"
        )