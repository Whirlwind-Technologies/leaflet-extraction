"""
Export Service Module.

This module provides comprehensive export functionality for leaflet data
in multiple formats (JSON, CSV, Excel) with configurable options.

Supports two export modes:
1. **Leaflet-level exports** (original): Export all products from a single leaflet.
2. **Product-level exports** (new): Export products across multiple leaflets
   with filtering, selection, and async job support for large result sets.

Example Usage:
    from app.services.export_service import ExportService

    service = ExportService(db)

    # Leaflet-level export
    excel_bytes = await service.export_excel(leaflet_id, user_id)

    # Product-level export
    from app.schemas.product_export import ProductExportRequest, ExportMode
    request = ProductExportRequest(format="csv", mode="all", image_storage="url")
    file_bytes = await service.export_products(request, org_id)
"""

import csv
import io
import json
import logging
from collections import defaultdict
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple, TYPE_CHECKING
from uuid import UUID

if TYPE_CHECKING:
    from app.schemas.product_export import ProductExportRequest

from sqlalchemy import func, select, and_
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.leaflet import Leaflet, LeafletPage
from app.models.product import Product, ReviewStatus
from app.utils.exceptions import NotFoundError, ValidationException

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# CSV column definitions for product-level exports
# ---------------------------------------------------------------------------
PRODUCT_EXPORT_CSV_COLUMNS = [
    "Leaflet ID",
    "Leaflet Name",
    "Page",
    "Brand",
    "Product Code",
    "Product Name",
    "Quantity",
    "Units",
    "Regular Price",
    "Discounted Price",
    "Discount %",
    "Currency",
    "Product ID",
    "Promotional Info",
    "Category",
    "Confidence",
    "Review Status",
    "Validation Passed",
    "Image URL",
    "Created At",
]

# Estimated bytes per product for file size estimation
SIZE_ESTIMATES_PER_PRODUCT = {
    ("csv", "none"): 500,
    ("csv", "url"): 600,
    ("csv", "base64"): 50_000,
    ("excel", "none"): 700,
    ("excel", "url"): 800,
    ("excel", "base64"): 50_000,
    ("json", "none"): 1_000,
    ("json", "url"): 1_200,
    ("json", "base64"): 50_000,
}


class ExportConfig:
    """Export configuration options."""
    
    def __init__(
        self,
        include_images: bool = True,
        image_format: str = "url",  # url, base64, both, none
        include_product_codes: bool = True,
        include_bounding_boxes: bool = True,
        include_confidence: bool = True,
        include_metadata: bool = True,
        include_validation: bool = False,
        columns: Optional[List[str]] = None,
        sort_by: str = "page_number",
        sort_order: str = "asc",
        filter_status: Optional[str] = None,
        min_confidence: Optional[float] = None,
    ):
        self.include_images = include_images
        self.image_format = image_format
        self.include_product_codes = include_product_codes
        self.include_bounding_boxes = include_bounding_boxes
        self.include_confidence = include_confidence
        self.include_metadata = include_metadata
        self.include_validation = include_validation
        self.columns = columns
        self.sort_by = sort_by
        self.sort_order = sort_order
        self.filter_status = filter_status
        self.min_confidence = min_confidence


class ExportService:
    """
    Service for exporting leaflet data in various formats.
    
    Supports JSON, CSV, and Excel exports with configurable options.
    """
    
    def __init__(self, db: AsyncSession):
        self.db = db
    
    async def get_leaflet_with_products(
        self,
        leaflet_id: str,
        user_id: UUID,
        config: Optional[ExportConfig] = None,
    ) -> Tuple[Leaflet, List[Product]]:
        """
        Get leaflet and products for export.
        
        Args:
            leaflet_id: Leaflet ID (human-readable or UUID)
            user_id: Owner user ID
            config: Export configuration
            
        Returns:
            Tuple of (Leaflet, list of Products)
        """
        config = config or ExportConfig()
        
        # Find leaflet
        query = select(Leaflet).where(Leaflet.user_id == user_id)
        
        try:
            uuid_id = UUID(leaflet_id)
            query = query.where(Leaflet.id == uuid_id)
        except ValueError:
            query = query.where(Leaflet.leaflet_id == leaflet_id)
        
        result = await self.db.execute(query)
        leaflet = result.scalar_one_or_none()
        
        if leaflet is None:
            raise NotFoundError("Leaflet", leaflet_id)
        
        # Get products
        products_query = select(Product).where(Product.leaflet_id == leaflet.id)
        
        # Apply filters
        if config.filter_status:
            products_query = products_query.where(
                Product.review_status == config.filter_status
            )
        
        if config.min_confidence is not None:
            products_query = products_query.where(
                Product.confidence >= config.min_confidence
            )
        
        # Apply sorting
        sort_column = getattr(Product, config.sort_by, Product.page_number)
        if config.sort_order == "desc":
            products_query = products_query.order_by(sort_column.desc())
        else:
            products_query = products_query.order_by(sort_column.asc())
        
        products_result = await self.db.execute(products_query)
        products = products_result.scalars().all()
        
        return leaflet, products
    
    def _product_to_dict(
        self,
        product: Product,
        config: ExportConfig,
        leaflet: Leaflet,
    ) -> Dict[str, Any]:
        """Convert product to dictionary for export."""
        data = {
            "page": product.page_number,
            "brand": product.brand,
            "product_name": product.product_name,
            "quantity": product.quantity,
            "units": product.units,
            "regular_price": product.regular_price,
            "discounted_price": product.discounted_price,
            "discount_percentage": product.discount_percentage,
            "currency": product.currency or leaflet.currency,
            "product_id": product.product_id,
            "promotional_info": product.promotional_info,
            "suggested_category": product.suggested_category,
            "category": product.category,
            "category_confidence": product.category_confidence,
        }
        
        if config.include_product_codes:
            data["product_code"] = product.product_code
        
        if config.include_bounding_boxes:
            data["bounding_box"] = {
                "x": product.bbox_x,
                "y": product.bbox_y,
                "width": product.bbox_width,
                "height": product.bbox_height,
            }
        
        if config.include_confidence:
            data["confidence"] = product.confidence
            data["field_confidence"] = product.field_confidence
        
        if config.include_validation:
            data["validation_passed"] = product.validation_passed
            data["validation_errors"] = product.validation_errors
            data["review_status"] = product.review_status.value if product.review_status else None
        
        if config.include_images and config.image_format != "none":
            image_data = {}
            
            if config.image_format in ["url", "both"] and product.image_url:
                image_data["url"] = product.image_url
            
            if config.image_format in ["base64", "both"] and product.image_base64:
                image_data["data"] = product.image_base64
            
            if image_data:
                image_data["format"] = product.image_format
                image_data["dimensions"] = {
                    "width": product.image_width,
                    "height": product.image_height,
                }
                data["image"] = image_data
        
        return data
    
    async def export_json(
        self,
        leaflet_id: str,
        user_id: UUID,
        config: Optional[ExportConfig] = None,
        pretty: bool = True,
    ) -> str:
        """
        Export leaflet data as JSON.
        
        Args:
            leaflet_id: Leaflet ID
            user_id: Owner user ID
            config: Export configuration
            pretty: Pretty print JSON
            
        Returns:
            JSON string
        """
        config = config or ExportConfig()
        leaflet, products = await self.get_leaflet_with_products(
            leaflet_id, user_id, config
        )
        
        export_data = {
            "leaflet_id": leaflet.leaflet_id,
            "exported_at": datetime.utcnow().isoformat(),
            "total_pages": leaflet.page_count,
            "total_products": len(products),
            "retailer": leaflet.retailer,
            "country": leaflet.country,
            "language": leaflet.language,
            "currency": leaflet.currency,
            "products": [
                self._product_to_dict(p, config, leaflet)
                for p in products
            ],
        }
        
        if config.include_metadata:
            export_data["metadata"] = {
                "processing_started_at": leaflet.processing_started_at.isoformat() if leaflet.processing_started_at else None,
                "processing_completed_at": leaflet.processing_completed_at.isoformat() if leaflet.processing_completed_at else None,
                "status": leaflet.status.value if leaflet.status else None,
            }
            
            export_data["quality_metrics"] = {
                "overall_confidence": leaflet.overall_confidence,
                "auto_approved_count": leaflet.auto_approved_count,
                "review_required_count": leaflet.review_required_count,
                "api_tokens_used": leaflet.api_tokens_used,
                "processing_cost": leaflet.processing_cost,
            }
        
        if pretty:
            return json.dumps(export_data, indent=2, default=str)
        return json.dumps(export_data, default=str)
    
    async def export_csv(
        self,
        leaflet_id: str,
        user_id: UUID,
        config: Optional[ExportConfig] = None,
    ) -> Tuple[str, str]:
        """
        Export leaflet data as CSV.
        
        Args:
            leaflet_id: Leaflet ID
            user_id: Owner user ID
            config: Export configuration
            
        Returns:
            Tuple of (CSV string, filename)
        """
        config = config or ExportConfig()
        leaflet, products = await self.get_leaflet_with_products(
            leaflet_id, user_id, config
        )
        
        output = io.StringIO()
        
        # Define columns
        columns = [
            "Leaflet_ID", "Page", "Brand", "Product_Name", "Quantity", "Units",
            "Regular_Price", "Discounted_Price", "Discount_Percentage", "Currency",
            "Product_ID", "Promotional_Info", "Category",
        ]
        
        if config.include_product_codes:
            columns.insert(3, "Product_Code")
        
        if config.include_bounding_boxes:
            columns.extend([
                "BoundingBox_X", "BoundingBox_Y",
                "BoundingBox_Width", "BoundingBox_Height"
            ])
        
        if config.include_confidence:
            columns.append("Confidence")
        
        if config.include_validation:
            columns.extend(["Review_Status", "Validation_Passed"])
        
        if config.include_images and config.image_format in ["url", "both"]:
            columns.append("Image_URL")
        
        if config.include_images and config.image_format in ["base64", "both"]:
            columns.append("Image_Base64")
        
        writer = csv.DictWriter(output, fieldnames=columns)
        writer.writeheader()
        
        for product in products:
            row = {
                "Leaflet_ID": leaflet.leaflet_id,
                "Page": product.page_number,
                "Brand": product.brand or "",
                "Product_Name": product.product_name,
                "Quantity": product.quantity or "",
                "Units": product.units or "",
                "Regular_Price": product.regular_price or "",
                "Discounted_Price": product.discounted_price or "",
                "Discount_Percentage": product.discount_percentage or "",
                "Currency": product.currency or leaflet.currency or "",
                "Product_ID": product.product_id or "",
                "Promotional_Info": product.promotional_info or "",
                "Category": product.category or "",
            }
            
            if config.include_product_codes:
                row["Product_Code"] = product.product_code or ""
            
            if config.include_bounding_boxes:
                row["BoundingBox_X"] = product.bbox_x
                row["BoundingBox_Y"] = product.bbox_y
                row["BoundingBox_Width"] = product.bbox_width
                row["BoundingBox_Height"] = product.bbox_height
            
            if config.include_confidence:
                row["Confidence"] = product.confidence or ""
            
            if config.include_validation:
                row["Review_Status"] = product.review_status.value if product.review_status else ""
                row["Validation_Passed"] = product.validation_passed
            
            if config.include_images and config.image_format in ["url", "both"]:
                row["Image_URL"] = product.image_url or ""
            
            if config.include_images and config.image_format in ["base64", "both"]:
                row["Image_Base64"] = product.image_base64 or ""
            
            writer.writerow(row)
        
        output.seek(0)
        filename = f"{leaflet.leaflet_id}_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
        
        return output.getvalue(), filename
    
    async def export_excel(
        self,
        leaflet_id: str,
        user_id: UUID,
        config: Optional[ExportConfig] = None,
    ) -> Tuple[bytes, str]:
        """
        Export leaflet data as Excel file.
        
        Args:
            leaflet_id: Leaflet ID
            user_id: Owner user ID
            config: Export configuration
            
        Returns:
            Tuple of (Excel bytes, filename)
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install it with: pip install openpyxl"
            )
        
        config = config or ExportConfig()
        leaflet, products = await self.get_leaflet_with_products(
            leaflet_id, user_id, config
        )
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # --- Summary Sheet ---
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Summary data
        summary_data = [
            ("Leaflet ID", leaflet.leaflet_id),
            ("Retailer", leaflet.retailer or "N/A"),
            ("Country", leaflet.country or "N/A"),
            ("Language", leaflet.language or "N/A"),
            ("Currency", leaflet.currency or "N/A"),
            ("Total Pages", leaflet.page_count),
            ("Total Products", len(products)),
            ("Auto Approved", leaflet.auto_approved_count or 0),
            ("Review Required", leaflet.review_required_count or 0),
            ("Overall Confidence", f"{(leaflet.overall_confidence or 0) * 100:.1f}%"),
            ("Processing Cost", f"${leaflet.processing_cost or 0:.4f}"),
            ("Exported At", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")),
        ]
        
        for row_idx, (label, value) in enumerate(summary_data, start=1):
            ws_summary.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws_summary.cell(row=row_idx, column=2, value=value)
        
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 30
        
        # --- Products Sheet ---
        ws_products = wb.create_sheet("Products")
        
        # Headers
        headers = [
            "Page", "Brand", "Product Name", "Quantity", "Units",
            "Regular Price", "Discounted Price", "Discount %", "Currency",
            "Product ID", "Promotional Info", "Category",
        ]
        
        if config.include_product_codes:
            headers.insert(2, "Product Code")
        
        if config.include_confidence:
            headers.append("Confidence")
        
        if config.include_validation:
            headers.extend(["Status", "Validated"])
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_products.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Write data
        for row_idx, product in enumerate(products, start=2):
            col_idx = 1
            
            ws_products.cell(row=row_idx, column=col_idx, value=product.page_number)
            col_idx += 1
            
            ws_products.cell(row=row_idx, column=col_idx, value=product.brand or "")
            col_idx += 1
            
            if config.include_product_codes:
                ws_products.cell(row=row_idx, column=col_idx, value=product.product_code or "")
                col_idx += 1
            
            ws_products.cell(row=row_idx, column=col_idx, value=product.product_name)
            col_idx += 1
            
            ws_products.cell(row=row_idx, column=col_idx, value=product.quantity or "")
            col_idx += 1
            
            ws_products.cell(row=row_idx, column=col_idx, value=product.units or "")
            col_idx += 1
            
            ws_products.cell(row=row_idx, column=col_idx, value=product.regular_price or "")
            col_idx += 1
            
            ws_products.cell(row=row_idx, column=col_idx, value=product.discounted_price or "")
            col_idx += 1
            
            ws_products.cell(row=row_idx, column=col_idx, value=product.discount_percentage or "")
            col_idx += 1
            
            ws_products.cell(row=row_idx, column=col_idx, value=product.currency or leaflet.currency or "")
            col_idx += 1
            
            ws_products.cell(row=row_idx, column=col_idx, value=product.product_id or "")
            col_idx += 1
            
            ws_products.cell(row=row_idx, column=col_idx, value=product.promotional_info or "")
            col_idx += 1

            ws_products.cell(row=row_idx, column=col_idx, value=product.category or "")
            col_idx += 1

            if config.include_confidence:
                conf_cell = ws_products.cell(
                    row=row_idx, column=col_idx,
                    value=f"{(product.confidence or 0) * 100:.1f}%"
                )
                # Color code confidence
                if product.confidence and product.confidence >= 0.9:
                    conf_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif product.confidence and product.confidence >= 0.75:
                    conf_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    conf_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                col_idx += 1
            
            if config.include_validation:
                ws_products.cell(
                    row=row_idx, column=col_idx,
                    value=product.review_status.value if product.review_status else ""
                )
                col_idx += 1
                
                ws_products.cell(
                    row=row_idx, column=col_idx,
                    value="Yes" if product.validation_passed else "No"
                )
                col_idx += 1
        
        # Auto-size columns
        for col_idx in range(1, len(headers) + 1):
            ws_products.column_dimensions[get_column_letter(col_idx)].width = 15
        
        # Product Name column wider
        name_col_idx = 4 if config.include_product_codes else 3
        ws_products.column_dimensions[get_column_letter(name_col_idx)].width = 40
        
        # Add auto-filter
        ws_products.auto_filter.ref = ws_products.dimensions
        
        # Freeze header row
        ws_products.freeze_panes = 'A2'
        
        # --- Stats Sheet ---
        ws_stats = wb.create_sheet("Statistics")
        
        # Calculate stats
        total_products = len(products)
        avg_confidence = sum(p.confidence or 0 for p in products) / max(total_products, 1)
        products_by_page = {}
        for p in products:
            products_by_page[p.page_number] = products_by_page.get(p.page_number, 0) + 1
        
        brands = {}
        for p in products:
            brand = p.brand or "Unknown"
            brands[brand] = brands.get(brand, 0) + 1
        
        # Write stats
        ws_stats.cell(row=1, column=1, value="Products by Page").font = Font(bold=True)
        row_idx = 2
        for page, count in sorted(products_by_page.items()):
            ws_stats.cell(row=row_idx, column=1, value=f"Page {page}")
            ws_stats.cell(row=row_idx, column=2, value=count)
            row_idx += 1
        
        row_idx += 1
        ws_stats.cell(row=row_idx, column=1, value="Products by Brand").font = Font(bold=True)
        row_idx += 1
        for brand, count in sorted(brands.items(), key=lambda x: x[1], reverse=True)[:10]:
            ws_stats.cell(row=row_idx, column=1, value=brand)
            ws_stats.cell(row=row_idx, column=2, value=count)
            row_idx += 1
        
        ws_stats.column_dimensions['A'].width = 25
        ws_stats.column_dimensions['B'].width = 10
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"{leaflet.leaflet_id}_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return output.getvalue(), filename
    
    async def export_bulk(
        self,
        leaflet_ids: List[str],
        user_id: UUID,
        format: str = "json",
        config: Optional[ExportConfig] = None,
    ) -> Dict[str, Any]:
        """
        Export multiple leaflets.
        
        Args:
            leaflet_ids: List of leaflet IDs
            user_id: Owner user ID
            format: Export format
            config: Export configuration
            
        Returns:
            Dictionary with export results
        """
        config = config or ExportConfig()
        results = {
            "exported_at": datetime.utcnow().isoformat(),
            "total_leaflets": len(leaflet_ids),
            "successful": 0,
            "failed": 0,
            "leaflets": [],
            "errors": [],
        }
        
        for leaflet_id in leaflet_ids:
            try:
                if format == "json":
                    data = json.loads(await self.export_json(leaflet_id, user_id, config))
                    results["leaflets"].append(data)
                else:
                    # For CSV/Excel, return metadata only
                    leaflet, products = await self.get_leaflet_with_products(
                        leaflet_id, user_id, config
                    )
                    results["leaflets"].append({
                        "leaflet_id": leaflet.leaflet_id,
                        "product_count": len(products),
                    })
                
                results["successful"] += 1
                
            except Exception as e:
                results["failed"] += 1
                results["errors"].append({
                    "leaflet_id": leaflet_id,
                    "error": str(e),
                })
        
        return results

    # ------------------------------------------------------------------
    # Product-level export methods (cross-leaflet)
    # ------------------------------------------------------------------

    async def count_products_for_export(
        self,
        request: "ProductExportRequest",
        organization_id: UUID,
    ) -> Tuple[int, int]:
        """
        Count products and distinct leaflets matching the export criteria.

        This executes an aggregate-only query (no row fetching) for fast
        preview responses.

        Args:
            request: Validated ProductExportRequest with mode and filters.
            organization_id: Organization UUID for data isolation.

        Returns:
            Tuple of (product_count, leaflet_count).
        """
        from app.schemas.product_export import ExportMode

        conditions = [Product.organization_id == organization_id]
        conditions = self._apply_export_mode_conditions(
            conditions, request, organization_id
        )

        count_query = select(
            func.count(Product.id).label("product_count"),
            func.count(func.distinct(Product.leaflet_id)).label("leaflet_count"),
        ).where(and_(*conditions))

        result = await self.db.execute(count_query)
        row = result.one()
        return row.product_count, row.leaflet_count

    def estimate_file_size(
        self,
        product_count: int,
        export_format: str,
        image_storage: str,
    ) -> str:
        """
        Estimate the export file size and return a human-readable string.

        Args:
            product_count: Number of products in the export.
            export_format: File format (csv, excel, json).
            image_storage: Image inclusion mode (url, base64, none).

        Returns:
            Human-readable size string (e.g. "1.2 MB").
        """
        per_product = SIZE_ESTIMATES_PER_PRODUCT.get(
            (export_format, image_storage),
            SIZE_ESTIMATES_PER_PRODUCT.get((export_format, "none"), 700),
        )
        total_bytes = product_count * per_product
        return _format_file_size(total_bytes)

    async def export_products(
        self,
        request: "ProductExportRequest",
        organization_id: UUID,
    ) -> io.BytesIO:
        """
        Export products across multiple leaflets to a file-like BytesIO object.

        Uses ``stream_products_for_export`` from database.py to fetch products
        in batches, then delegates to the appropriate format writer.

        Args:
            request: Validated ProductExportRequest with mode, format,
                     filters, and image_storage preference.
            organization_id: Organization UUID for data isolation.

        Returns:
            BytesIO containing the complete export file bytes.

        Raises:
            ValidationException: If request parameters are invalid.
        """
        from app.schemas.product_export import ExportFormat
        from app.utils.database import stream_products_for_export

        # Build filter dict consumed by stream_products_for_export. This
        # resolves human-readable leaflet IDs to UUIDs against the calling
        # organization so the export uses the same scope as the preview.
        filters = await self._build_stream_filters(request, organization_id)

        # Collect all products (we need them all for Excel summary sheets
        # and JSON structure; streaming is used for memory-efficient DB fetch)
        all_products: List[Product] = []
        async for batch in stream_products_for_export(
            self.db, filters, organization_id, batch_size=500
        ):
            all_products.extend(batch)

        # Pre-fetch leaflet metadata for the leaflet columns
        leaflet_map = await self._build_leaflet_map(all_products)

        # Determine image inclusion mode
        image_storage = request.image_storage.value

        if request.format == ExportFormat.CSV:
            return self._write_products_csv(
                all_products, leaflet_map, image_storage
            )
        elif request.format == ExportFormat.EXCEL:
            return self._write_products_excel(
                all_products, leaflet_map, image_storage, request
            )
        elif request.format == ExportFormat.JSON:
            return self._write_products_json(
                all_products, leaflet_map, image_storage
            )
        else:
            raise ValidationException(
                [{"field": "format", "message": f"Unsupported format: {request.format}"}]
            )

    # ------------------------------------------------------------------
    # Internal helpers: query building
    # ------------------------------------------------------------------

    def _apply_export_mode_conditions(
        self,
        conditions: list,
        request: "ProductExportRequest",
        organization_id: Optional[UUID] = None,
    ) -> list:
        """
        Append WHERE conditions based on the export mode and its filters.

        Args:
            conditions: Existing list of SQLAlchemy conditions.
            request: Validated ProductExportRequest.
            organization_id: Organization UUID for scoping leaflet lookups.

        Returns:
            Updated conditions list (same object, mutated in place).
        """
        from app.schemas.product_export import ExportMode

        if request.mode == ExportMode.FILTERED and request.filters:
            filters = request.filters
            if filters.search:
                conditions.append(
                    Product.product_name.ilike(f"%{filters.search}%")
                )
            if filters.review_status:
                enum_values = [ReviewStatus(s) for s in filters.review_status]
                conditions.append(Product.review_status.in_(enum_values))
            if filters.leaflet_id:
                leaflet_id_val = filters.leaflet_id
                try:
                    leaflet_uuid = UUID(leaflet_id_val)
                    conditions.append(Product.leaflet_id == leaflet_uuid)
                except ValueError:
                    # Human-readable ID -- use a subquery
                    sub_conditions = [Leaflet.leaflet_id == leaflet_id_val]
                    if organization_id is not None:
                        sub_conditions.append(
                            Leaflet.organization_id == organization_id
                        )
                    leaflet_sub = select(Leaflet.id).where(
                        and_(*sub_conditions)
                    )
                    conditions.append(Product.leaflet_id.in_(leaflet_sub))
            if filters.category:
                conditions.append(Product.category == filters.category)
            if filters.brand:
                conditions.append(Product.brand.ilike(f"%{filters.brand}%"))
            if filters.min_confidence is not None:
                conditions.append(
                    Product.confidence >= filters.min_confidence
                )
            if filters.page_number is not None:
                conditions.append(
                    Product.page_number == filters.page_number
                )
            if filters.validation_passed is not None:
                conditions.append(
                    Product.validation_passed == filters.validation_passed
                )

        elif request.mode == ExportMode.SELECTED and request.product_ids:
            conditions.append(Product.id.in_(request.product_ids))

        elif request.mode == ExportMode.REVIEW_QUEUE:
            conditions.append(
                Product.review_status.in_([
                    ReviewStatus.PENDING,
                    ReviewStatus.NEEDS_CORRECTION,
                ])
            )
            if (
                request.review_queue_filters
                and request.review_queue_filters.leaflet_id
            ):
                leaflet_id_val = request.review_queue_filters.leaflet_id
                try:
                    leaflet_uuid = UUID(leaflet_id_val)
                    conditions.append(Product.leaflet_id == leaflet_uuid)
                except ValueError:
                    sub_conditions = [Leaflet.leaflet_id == leaflet_id_val]
                    if organization_id is not None:
                        sub_conditions.append(
                            Leaflet.organization_id == organization_id
                        )
                    leaflet_sub = select(Leaflet.id).where(
                        and_(*sub_conditions)
                    )
                    conditions.append(Product.leaflet_id.in_(leaflet_sub))

        # mode == ExportMode.ALL requires no additional conditions

        return conditions

    async def _resolve_leaflet_id(
        self,
        leaflet_id_value: str,
        organization_id: UUID,
    ) -> Optional[UUID]:
        """
        Resolve a leaflet identifier (either a UUID or a human-readable
        ``LEAF_YYYY_XXXXXX`` string) to a UUID scoped to the given org.

        Returns the UUID, or None if no matching leaflet exists in this
        organization. Returning None causes the caller to apply an
        impossible filter so the export is empty, matching the
        count_products preview behaviour.
        """
        try:
            return UUID(leaflet_id_value)
        except ValueError:
            pass

        result = await self.db.execute(
            select(Leaflet.id).where(
                and_(
                    Leaflet.leaflet_id == leaflet_id_value,
                    Leaflet.organization_id == organization_id,
                )
            )
        )
        row = result.first()
        return row[0] if row else None

    async def _build_stream_filters(
        self,
        request: "ProductExportRequest",
        organization_id: UUID,
    ) -> dict:
        """
        Convert a ProductExportRequest into the flat dict expected by
        ``stream_products_for_export()``.

        Resolves human-readable leaflet IDs against the calling
        organization so the export uses the same filter semantics
        as ``count_products_for_export`` (the preview endpoint).
        Previously, human-readable IDs were silently dropped here,
        making the export contain more rows than the preview claimed.
        """
        from app.schemas.product_export import ExportMode

        filters: Dict[str, Any] = {}

        # Sentinel UUID for "no leaflet exists" — leaks zero products
        # instead of all products when a filter cannot be resolved.
        EMPTY_UUID = UUID("00000000-0000-0000-0000-000000000000")

        if request.mode == ExportMode.FILTERED and request.filters:
            f = request.filters
            if f.search:
                filters["search"] = f.search
            if f.review_status:
                filters["review_statuses"] = f.review_status
            if f.leaflet_id:
                resolved = await self._resolve_leaflet_id(
                    f.leaflet_id, organization_id
                )
                filters["leaflet_id"] = resolved if resolved is not None else EMPTY_UUID
            if f.category:
                filters["category"] = f.category
            if f.brand:
                filters["brand"] = f.brand
            if f.min_confidence is not None:
                filters["min_confidence"] = f.min_confidence
            if f.validation_passed is not None:
                filters["validation_passed"] = f.validation_passed
            filters["sort_by"] = f.sort_by
            filters["sort_order"] = f.sort_order

        elif request.mode == ExportMode.SELECTED and request.product_ids:
            filters["product_ids"] = request.product_ids

        elif request.mode == ExportMode.REVIEW_QUEUE:
            filters["review_statuses"] = [
                ReviewStatus.PENDING.value,
                ReviewStatus.NEEDS_CORRECTION.value,
            ]
            if (
                request.review_queue_filters
                and request.review_queue_filters.leaflet_id
            ):
                resolved = await self._resolve_leaflet_id(
                    request.review_queue_filters.leaflet_id, organization_id
                )
                filters["leaflet_id"] = resolved if resolved is not None else EMPTY_UUID

        return filters

    async def _build_leaflet_map(
        self,
        products: List[Product],
    ) -> Dict[UUID, Leaflet]:
        """
        Fetch Leaflet metadata for every distinct leaflet_id in the products.

        Args:
            products: List of Product ORM instances.

        Returns:
            Dict mapping leaflet UUID -> Leaflet instance.
        """
        leaflet_ids = list({p.leaflet_id for p in products})
        if not leaflet_ids:
            return {}

        result = await self.db.execute(
            select(Leaflet).where(Leaflet.id.in_(leaflet_ids))
        )
        leaflets = result.scalars().all()
        return {lf.id: lf for lf in leaflets}

    # ------------------------------------------------------------------
    # Internal helpers: format writers (product-level)
    # ------------------------------------------------------------------

    def _write_products_csv(
        self,
        products: List[Product],
        leaflet_map: Dict[UUID, Leaflet],
        image_storage: str,
    ) -> io.BytesIO:
        """
        Write products to CSV format.

        Args:
            products: List of Product ORM instances.
            leaflet_map: Dict mapping leaflet UUID -> Leaflet.
            image_storage: How to include images (url, base64, none).

        Returns:
            BytesIO containing UTF-8 encoded CSV.
        """
        output = io.StringIO()
        columns = list(PRODUCT_EXPORT_CSV_COLUMNS)

        # Adjust image column based on image_storage preference
        if image_storage == "none":
            columns.remove("Image URL")
        elif image_storage == "base64":
            idx = columns.index("Image URL")
            columns[idx] = "Image Base64"

        writer = csv.DictWriter(output, fieldnames=columns)
        writer.writeheader()

        for product in products:
            leaflet = leaflet_map.get(product.leaflet_id)
            row = self._product_to_csv_row(product, leaflet, image_storage)
            writer.writerow(row)

        csv_bytes = output.getvalue().encode("utf-8")
        return io.BytesIO(csv_bytes)

    def _product_to_csv_row(
        self,
        product: Product,
        leaflet: Optional[Leaflet],
        image_storage: str,
    ) -> Dict[str, Any]:
        """
        Convert a single Product to a CSV row dict.

        Uses the PRODUCT_EXPORT_CSV_COLUMNS layout.

        Args:
            product: Product ORM instance.
            leaflet: Parent Leaflet (may be None if deleted).
            image_storage: Image inclusion mode.

        Returns:
            Dict keyed by column name.
        """
        leaflet_id_str = leaflet.leaflet_id if leaflet else ""
        leaflet_name = leaflet.filename if leaflet else ""
        currency = product.currency or (leaflet.currency if leaflet else "") or ""

        row: Dict[str, Any] = {
            "Leaflet ID": leaflet_id_str,
            "Leaflet Name": leaflet_name,
            "Page": product.page_number,
            "Brand": product.brand or "",
            "Product Code": product.product_code or "",
            "Product Name": product.product_name or "",
            "Quantity": product.quantity if product.quantity is not None else "",
            "Units": product.units or "",
            "Regular Price": product.regular_price if product.regular_price is not None else "",
            "Discounted Price": product.discounted_price if product.discounted_price is not None else "",
            "Discount %": product.discount_percentage if product.discount_percentage is not None else "",
            "Currency": currency,
            "Product ID": product.product_id or "",
            "Promotional Info": product.promotional_info or "",
            "Category": product.category or "",
            "Confidence": product.confidence if product.confidence is not None else "",
            "Review Status": product.review_status.value if product.review_status else "",
            "Validation Passed": "Yes" if product.validation_passed else "No",
            "Created At": product.created_at.isoformat() if product.created_at else "",
        }

        if image_storage == "url":
            row["Image URL"] = product.image_url or ""
        elif image_storage == "base64":
            row["Image Base64"] = product.image_base64 or ""
        # none: column not present

        return row

    def _write_products_excel(
        self,
        products: List[Product],
        leaflet_map: Dict[UUID, Leaflet],
        image_storage: str,
        request: "ProductExportRequest",
    ) -> io.BytesIO:
        """
        Write products to Excel format with three sheets.

        Sheet 1 (Products): All product data rows.
        Sheet 2 (Summary): Counts by status, avg confidence, leaflet breakdown.
        Sheet 3 (Export Metadata): Filters applied, date, user info.

        Args:
            products: List of Product ORM instances.
            leaflet_map: Dict mapping leaflet UUID -> Leaflet.
            image_storage: Image inclusion mode.
            request: Original ProductExportRequest for metadata sheet.

        Returns:
            BytesIO containing the Excel workbook.
        """
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"
        )
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # ---- Sheet 1: Products ----
        ws_products = wb.active
        ws_products.title = "Products"

        headers = [
            "Leaflet ID",
            "Leaflet Name",
            "Page",
            "Brand",
            "Product Code",
            "Product Name",
            "Quantity",
            "Units",
            "Regular Price",
            "Discounted Price",
            "Discount %",
            "Currency",
            "Product ID",
            "Promotional Info",
            "Category",
            "Confidence",
            "Review Status",
            "Validation Passed",
        ]

        if image_storage == "url":
            headers.append("Image URL")
        elif image_storage == "base64":
            headers.append("Image Base64")

        headers.append("Created At")

        # Write header row
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_products.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        # Write data rows
        for row_idx, product in enumerate(products, start=2):
            leaflet = leaflet_map.get(product.leaflet_id)
            col = 1

            ws_products.cell(row=row_idx, column=col, value=leaflet.leaflet_id if leaflet else "")
            col += 1
            ws_products.cell(row=row_idx, column=col, value=leaflet.filename if leaflet else "")
            col += 1
            ws_products.cell(row=row_idx, column=col, value=product.page_number)
            col += 1
            ws_products.cell(row=row_idx, column=col, value=product.brand or "")
            col += 1
            ws_products.cell(row=row_idx, column=col, value=product.product_code or "")
            col += 1
            ws_products.cell(row=row_idx, column=col, value=product.product_name or "")
            col += 1
            ws_products.cell(row=row_idx, column=col, value=product.quantity if product.quantity is not None else "")
            col += 1
            ws_products.cell(row=row_idx, column=col, value=product.units or "")
            col += 1
            ws_products.cell(row=row_idx, column=col, value=product.regular_price if product.regular_price is not None else "")
            col += 1
            ws_products.cell(row=row_idx, column=col, value=product.discounted_price if product.discounted_price is not None else "")
            col += 1
            ws_products.cell(row=row_idx, column=col, value=product.discount_percentage if product.discount_percentage is not None else "")
            col += 1
            ws_products.cell(row=row_idx, column=col, value=product.currency or (leaflet.currency if leaflet else "") or "")
            col += 1
            ws_products.cell(row=row_idx, column=col, value=product.product_id or "")
            col += 1
            ws_products.cell(row=row_idx, column=col, value=product.promotional_info or "")
            col += 1
            ws_products.cell(row=row_idx, column=col, value=product.category or "")
            col += 1

            # Confidence with color coding
            conf_cell = ws_products.cell(
                row=row_idx,
                column=col,
                value=product.confidence if product.confidence is not None else "",
            )
            if product.confidence is not None:
                if product.confidence >= 0.9:
                    conf_cell.fill = PatternFill(
                        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                    )
                elif product.confidence >= 0.75:
                    conf_cell.fill = PatternFill(
                        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                    )
                else:
                    conf_cell.fill = PatternFill(
                        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                    )
            col += 1

            ws_products.cell(
                row=row_idx,
                column=col,
                value=product.review_status.value if product.review_status else "",
            )
            col += 1
            ws_products.cell(
                row=row_idx,
                column=col,
                value="Yes" if product.validation_passed else "No",
            )
            col += 1

            if image_storage == "url":
                ws_products.cell(row=row_idx, column=col, value=product.image_url or "")
                col += 1
            elif image_storage == "base64":
                ws_products.cell(row=row_idx, column=col, value=product.image_base64 or "")
                col += 1

            ws_products.cell(
                row=row_idx,
                column=col,
                value=product.created_at.isoformat() if product.created_at else "",
            )

        # Column widths
        for col_idx in range(1, len(headers) + 1):
            ws_products.column_dimensions[get_column_letter(col_idx)].width = 15
        # Product Name wider
        ws_products.column_dimensions[get_column_letter(6)].width = 40
        ws_products.auto_filter.ref = ws_products.dimensions
        ws_products.freeze_panes = "A2"

        # ---- Sheet 2: Summary ----
        ws_summary = wb.create_sheet("Summary")

        # Status counts
        status_counts: Dict[str, int] = defaultdict(int)
        leaflet_product_counts: Dict[str, int] = defaultdict(int)
        total_confidence = 0.0
        confidence_count = 0

        for p in products:
            status_val = p.review_status.value if p.review_status else "unknown"
            status_counts[status_val] += 1

            lf = leaflet_map.get(p.leaflet_id)
            lf_label = lf.leaflet_id if lf else str(p.leaflet_id)[:8]
            leaflet_product_counts[lf_label] += 1

            if p.confidence is not None:
                total_confidence += p.confidence
                confidence_count += 1

        avg_confidence = total_confidence / max(confidence_count, 1)

        row = 1
        ws_summary.cell(row=row, column=1, value="Products by Status").font = Font(bold=True)
        row += 1
        for status_name, count in sorted(status_counts.items()):
            ws_summary.cell(row=row, column=1, value=status_name)
            ws_summary.cell(row=row, column=2, value=count)
            row += 1

        row += 1
        ws_summary.cell(row=row, column=1, value="Overall Statistics").font = Font(bold=True)
        row += 1
        ws_summary.cell(row=row, column=1, value="Total Products")
        ws_summary.cell(row=row, column=2, value=len(products))
        row += 1
        ws_summary.cell(row=row, column=1, value="Distinct Leaflets")
        ws_summary.cell(row=row, column=2, value=len(leaflet_product_counts))
        row += 1
        ws_summary.cell(row=row, column=1, value="Average Confidence")
        ws_summary.cell(row=row, column=2, value=f"{avg_confidence * 100:.1f}%")
        row += 1

        row += 1
        ws_summary.cell(row=row, column=1, value="Products by Leaflet").font = Font(bold=True)
        row += 1
        for lf_label, count in sorted(
            leaflet_product_counts.items(), key=lambda x: x[1], reverse=True
        ):
            ws_summary.cell(row=row, column=1, value=lf_label)
            ws_summary.cell(row=row, column=2, value=count)
            row += 1

        ws_summary.column_dimensions["A"].width = 30
        ws_summary.column_dimensions["B"].width = 15

        # ---- Sheet 3: Export Metadata ----
        ws_meta = wb.create_sheet("Export Metadata")
        meta_data = [
            ("Export Date", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")),
            ("Export Mode", request.mode.value),
            ("Export Format", request.format.value),
            ("Image Storage", request.image_storage.value),
            ("Total Products", len(products)),
            ("Distinct Leaflets", len(leaflet_product_counts)),
        ]

        if request.filters:
            f = request.filters
            if f.search:
                meta_data.append(("Filter: Search", f.search))
            if f.review_status:
                meta_data.append(("Filter: Review Status", ", ".join(f.review_status)))
            if f.leaflet_id:
                meta_data.append(("Filter: Leaflet ID", f.leaflet_id))
            if f.category:
                meta_data.append(("Filter: Category", f.category))
            if f.brand:
                meta_data.append(("Filter: Brand", f.brand))
            if f.min_confidence is not None:
                meta_data.append(("Filter: Min Confidence", str(f.min_confidence)))
            if f.validation_passed is not None:
                meta_data.append(("Filter: Validation Passed", str(f.validation_passed)))
            meta_data.append(("Sort By", f.sort_by))
            meta_data.append(("Sort Order", f.sort_order))

        for row_idx, (label, value) in enumerate(meta_data, start=1):
            ws_meta.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws_meta.cell(row=row_idx, column=2, value=value)

        ws_meta.column_dimensions["A"].width = 25
        ws_meta.column_dimensions["B"].width = 50

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _write_products_json(
        self,
        products: List[Product],
        leaflet_map: Dict[UUID, Leaflet],
        image_storage: str,
    ) -> io.BytesIO:
        """
        Write products to JSON format.

        Args:
            products: List of Product ORM instances.
            leaflet_map: Dict mapping leaflet UUID -> Leaflet.
            image_storage: Image inclusion mode.

        Returns:
            BytesIO containing UTF-8 encoded JSON.
        """
        export_data = {
            "exported_at": datetime.utcnow().isoformat(),
            "total_products": len(products),
            "total_leaflets": len({p.leaflet_id for p in products}),
            "products": [],
        }

        for product in products:
            leaflet = leaflet_map.get(product.leaflet_id)
            product_dict: Dict[str, Any] = {
                "id": str(product.id),
                "leaflet_id": leaflet.leaflet_id if leaflet else str(product.leaflet_id),
                "leaflet_name": leaflet.filename if leaflet else None,
                "retailer": leaflet.retailer if leaflet else None,
                "country": leaflet.country if leaflet else None,
                "page_number": product.page_number,
                "brand": product.brand,
                "product_code": product.product_code,
                "product_name": product.product_name,
                "quantity": product.quantity,
                "units": product.units,
                "size": product.size,
                "regular_price": product.regular_price,
                "discounted_price": product.discounted_price,
                "discount_percentage": product.discount_percentage,
                "currency": product.currency or (leaflet.currency if leaflet else None),
                "product_id": product.product_id,
                "promotional_info": product.promotional_info,
                "category": product.category,
                "confidence": product.confidence,
                "review_status": (
                    product.review_status.value if product.review_status else None
                ),
                "validation_passed": product.validation_passed,
                "created_at": (
                    product.created_at.isoformat() if product.created_at else None
                ),
            }

            # Image data
            if image_storage == "url" and product.image_url:
                product_dict["image_url"] = product.image_url
            elif image_storage == "base64" and product.image_base64:
                product_dict["image_base64"] = product.image_base64

            export_data["products"].append(product_dict)

        json_bytes = json.dumps(export_data, indent=2, default=str).encode("utf-8")
        return io.BytesIO(json_bytes)


# ---------------------------------------------------------------------------
# Module-level utility functions
# ---------------------------------------------------------------------------

def _format_file_size(size_bytes: int) -> str:
    """
    Format a byte count as a human-readable string.

    Args:
        size_bytes: File size in bytes.

    Returns:
        Formatted string (e.g. "1.2 MB", "450 KB", "12 B").
    """
    if size_bytes < 1024:
        return f"{size_bytes} B"
    elif size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"
    elif size_bytes < 1024 * 1024 * 1024:
        return f"{size_bytes / (1024 * 1024):.1f} MB"
    else:
        return f"{size_bytes / (1024 * 1024 * 1024):.1f} GB"


def estimate_export_size(
    product_count: int,
    export_format: str,
    image_storage: str,
) -> str:
    """
    Module-level convenience wrapper for file size estimation.

    Useful from Celery tasks that do not instantiate ExportService.

    Args:
        product_count: Number of products.
        export_format: File format (csv, excel, json).
        image_storage: Image mode (url, base64, none).

    Returns:
        Human-readable size string.
    """
    per_product = SIZE_ESTIMATES_PER_PRODUCT.get(
        (export_format, image_storage),
        SIZE_ESTIMATES_PER_PRODUCT.get((export_format, "none"), 700),
    )
    return _format_file_size(product_count * per_product)